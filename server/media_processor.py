"""
Media file processor for Smart TV Media Player
Handles file processing: thumbnails, XLSX to PNG conversion
"""
import os
import uuid
import json
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


UPLOAD_FOLDER = None
THUMBNAIL_FOLDER = None
ALLOWED_EXTENSIONS = {'mp4', 'jpg', 'jpeg', 'xlsx', 'avi', 'mov'}


def init_processor(upload_folder, thumbnail_folder):
    """Initialize processor with storage paths"""
    global UPLOAD_FOLDER, THUMBNAIL_FOLDER
    UPLOAD_FOLDER = upload_folder
    THUMBNAIL_FOLDER = thumbnail_folder
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    os.makedirs(THUMBNAIL_FOLDER, exist_ok=True)


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def get_file_type(filename):
    """Get file type from extension"""
    ext = filename.rsplit('.', 1)[1].lower()
    return ext


def get_display_type(file_type):
    """Determine display type based on file type"""
    if file_type in ('mp4', 'avi', 'mov'):
        return 'video'
    return 'image'


def generate_unique_filename(original_filename):
    """Generate a unique filename preserving extension"""
    ext = original_filename.rsplit('.', 1)[1].lower()
    unique_name = f"{uuid.uuid4().hex}.{ext}"
    return unique_name


def transcode_to_mp4(input_path, output_path):
    """Transcode input video to MP4 (H.264 / AAC) using FFmpeg"""
    import subprocess
    
    try:
        import static_ffmpeg
        static_ffmpeg.add_paths()
    except ImportError:
        print("[transcode] static-ffmpeg not found, trying system PATH...")
        
    cmd = [
        'ffmpeg', '-y',
        '-i', input_path,
        '-c:v', 'libx264',
        '-preset', 'fast',
        '-crf', '23',
        '-c:a', 'aac',
        '-b:a', '128k',
        '-pix_fmt', 'yuv420p',
        output_path
    ]
    
    try:
        print(f"[transcode] Converting {input_path} to {output_path}...")
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=120)
        if result.returncode == 0:
            print("[transcode] Successful conversion.")
            return True
        else:
            print(f"[transcode] FFmpeg error {result.returncode}: {result.stderr}")
            return False
    except Exception as e:
        print(f"[transcode] Exception during conversion: {e}")
        return False


def get_video_duration(filepath):
    """Get video duration in seconds using ffprobe"""
    import subprocess
    
    try:
        import static_ffmpeg
        static_ffmpeg.add_paths()
    except ImportError:
        pass
        
    cmd = [
        'ffprobe', '-v', 'error',
        '-show_entries', 'format=duration',
        '-of', 'default=noprint_wrappers=1:nokey=1',
        filepath
    ]
    
    try:
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=10)
        if result.returncode == 0:
            duration = float(result.stdout.strip())
            return int(round(duration))
    except Exception as e:
        print(f"[duration] Error getting duration for {filepath}: {e}")
        
    return None


def save_uploaded_file(file_storage):
    """
    Save an uploaded file and process it.
    Returns dict with file info or None on error.
    """
    if not file_storage or not file_storage.filename:
        return None

    original_name = file_storage.filename
    if not allowed_file(original_name):
        return None

    file_type = get_file_type(original_name)
    display_type = get_display_type(file_type)
    
    # If it is avi/mov, we will save to a temp file, transcode to mp4, and delete temp
    if file_type in ('avi', 'mov'):
        temp_filename = f"temp_{uuid.uuid4().hex}.{file_type}"
        temp_filepath = os.path.join(UPLOAD_FOLDER, temp_filename)
        file_storage.save(temp_filepath)
        
        filename = f"{uuid.uuid4().hex}.mp4"
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        
        success = transcode_to_mp4(temp_filepath, filepath)
        
        # Clean up temp file
        try:
            if os.path.exists(temp_filepath):
                os.remove(temp_filepath)
        except Exception:
            pass
            
        if not success:
            return None
            
        # Update type parameters, since it's now a converted mp4
        file_type = 'mp4'
    else:
        filename = generate_unique_filename(original_name)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file_storage.save(filepath)

    file_size = os.path.getsize(filepath)

    # Generate thumbnail and check duration
    thumbnail = None
    converted_files = []
    duration = None

    if file_type in ('jpg', 'jpeg'):
        thumbnail = generate_image_thumbnail(filepath, filename)
    elif file_type == 'mp4':
        thumbnail = generate_video_thumbnail(filepath, filename)
        duration = get_video_duration(filepath)
    elif file_type == 'xlsx':
        converted_files = convert_xlsx_to_images(filepath, filename)
        if converted_files:
            # Use first converted image as thumbnail
            first_img = converted_files[0]
            thumbnail = generate_image_thumbnail(
                os.path.join(UPLOAD_FOLDER, first_img),
                first_img
            )

    return {
        'filename': filename,
        'original_name': original_name,
        'file_type': file_type,
        'display_type': display_type,
        'file_size': file_size,
        'thumbnail': thumbnail,
        'converted_files': json.dumps(converted_files) if converted_files else None,
        'duration': duration
    }


def generate_image_thumbnail(filepath, filename, size=(300, 200)):
    """Generate thumbnail for an image file"""
    try:
        thumb_name = f"thumb_{os.path.splitext(filename)[0]}.jpg"
        thumb_path = os.path.join(THUMBNAIL_FOLDER, thumb_name)

        with Image.open(filepath) as img:
            img = img.convert('RGB')
            img.thumbnail(size, Image.Resampling.LANCZOS)
            img.save(thumb_path, 'JPEG', quality=85)

        return thumb_name
    except Exception as e:
        print(f"Error generating thumbnail for {filename}: {e}")
        return None


def generate_video_thumbnail(filepath, filename):
    """Generate placeholder thumbnail for video files"""
    try:
        thumb_name = f"thumb_{os.path.splitext(filename)[0]}.jpg"
        thumb_path = os.path.join(THUMBNAIL_FOLDER, thumb_name)

        # Create a styled placeholder thumbnail for video
        img = Image.new('RGB', (300, 200), color=(30, 30, 46))
        draw = ImageDraw.Draw(img)

        # Draw play button triangle
        center_x, center_y = 150, 90
        triangle_size = 30
        play_points = [
            (center_x - triangle_size // 2, center_y - triangle_size),
            (center_x - triangle_size // 2, center_y + triangle_size),
            (center_x + triangle_size, center_y)
        ]
        draw.polygon(play_points, fill=(137, 180, 250))

        # Draw circle around play button
        draw.ellipse(
            [center_x - 40, center_y - 40, center_x + 40, center_y + 40],
            outline=(137, 180, 250), width=3
        )

        # Add text
        try:
            font = ImageFont.truetype("arial.ttf", 14)
        except (OSError, IOError):
            font = ImageFont.load_default()

        text = "MP4 Video"
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        draw.text(((300 - text_width) // 2, 160), text, fill=(166, 173, 200), font=font)

        img.save(thumb_path, 'JPEG', quality=85)
        return thumb_name
    except Exception as e:
        print(f"Error generating video thumbnail for {filename}: {e}")
        return None


def convert_xlsx_to_images(filepath, filename):
    """
    Convert XLSX file sheets to PNG images.
    Each sheet becomes a separate image.
    """
    converted = []
    try:
        wb = load_workbook(filepath, data_only=True)

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]

            # Calculate dimensions
            max_row = min(ws.max_row or 1, 50)  # Limit to 50 rows
            max_col = min(ws.max_column or 1, 20)  # Limit to 20 columns

            if max_row == 0 or max_col == 0:
                continue

            # Cell dimensions
            cell_width = 150
            cell_height = 35
            header_height = 50
            padding = 20

            img_width = padding * 2 + max_col * cell_width
            img_height = padding + header_height + max_row * cell_height + padding

            # Ensure minimum size
            img_width = max(img_width, 800)
            img_height = max(img_height, 400)

            # Create image with dark background
            img = Image.new('RGB', (img_width, img_height), color=(30, 30, 46))
            draw = ImageDraw.Draw(img)

            try:
                font = ImageFont.truetype("arial.ttf", 13)
                header_font = ImageFont.truetype("arial.ttf", 16)
            except (OSError, IOError):
                font = ImageFont.load_default()
                header_font = font

            # Draw sheet title
            title = f"📊 {sheet_name}"
            draw.text((padding, 12), title, fill=(205, 214, 244), font=header_font)

            # Draw table
            start_y = header_height

            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = str(cell.value) if cell.value is not None else ''

                    x = padding + (col_idx - 1) * cell_width
                    y = start_y + (row_idx - 1) * cell_height

                    # Cell background
                    if row_idx == 1:
                        bg_color = (49, 50, 68)  # Header row
                        text_color = (137, 180, 250)
                    elif row_idx % 2 == 0:
                        bg_color = (35, 35, 52)
                        text_color = (205, 214, 244)
                    else:
                        bg_color = (30, 30, 46)
                        text_color = (205, 214, 244)

                    # Draw cell
                    draw.rectangle(
                        [x, y, x + cell_width, y + cell_height],
                        fill=bg_color, outline=(69, 71, 90)
                    )

                    # Truncate text if too long
                    display_text = value[:18] + '...' if len(value) > 20 else value
                    draw.text(
                        (x + 5, y + 8),
                        display_text,
                        fill=text_color,
                        font=font if row_idx > 1 else header_font
                    )

            # Save converted image
            conv_name = f"xlsx_{os.path.splitext(filename)[0]}_sheet{sheet_idx}.png"
            conv_path = os.path.join(UPLOAD_FOLDER, conv_name)
            img.save(conv_path, 'PNG')
            converted.append(conv_name)

        wb.close()
    except Exception as e:
        print(f"Error converting XLSX {filename}: {e}")

    return converted


def delete_media_files(filename, thumbnail, converted_files_json):
    """Delete all files associated with a media entry"""
    # Delete main file
    main_path = os.path.join(UPLOAD_FOLDER, filename)
    if os.path.exists(main_path):
        os.remove(main_path)

    # Delete thumbnail
    if thumbnail:
        thumb_path = os.path.join(THUMBNAIL_FOLDER, thumbnail)
        if os.path.exists(thumb_path):
            os.remove(thumb_path)

    # Delete converted files (XLSX)
    if converted_files_json:
        try:
            converted = json.loads(converted_files_json)
            for conv_file in converted:
                conv_path = os.path.join(UPLOAD_FOLDER, conv_file)
                if os.path.exists(conv_path):
                    os.remove(conv_path)
        except (json.JSONDecodeError, TypeError):
            pass
